import streamlit as st
import requests
import pandas as pd
import time
import os
import json
import math
from urllib.parse import quote
from datetime import datetime
from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

st.set_page_config(page_title="Selltic Scraper", page_icon="🔍", layout="wide")

# ── Selltic theming (białe karty / jasnoszare tło / fiolet) ───────────────────

SELLTIC_CSS = """
<style>
@import url('https://fonts.googleapis.com/css2?family=Inter:wght@400;500;600;700&display=swap');

:root {
    --sl-accent: #6C5CE7;
    --sl-accent-dark: #5A4BD1;
    --sl-bg: #F6F7F9;
    --sl-surface: #FFFFFF;
    --sl-border: #ECEEF3;
    --sl-text: #1F2430;
    --sl-text-muted: #767E8C;
}

html, body, [class*="css"] {
    font-family: 'Inter', 'Segoe UI', sans-serif;
    color: var(--sl-text);
}

.stApp {
    background-color: var(--sl-bg);
}

h1, h2, h3, h4 {
    font-weight: 700 !important;
    color: var(--sl-text);
    letter-spacing: -0.01em;
}

h1 {
    font-size: 1.7rem !important;
}

section[data-testid="stSidebar"] {
    background-color: var(--sl-surface);
    border-right: 1px solid var(--sl-border);
    box-shadow: none;
}

.sl-logo-row {
    display: flex;
    align-items: center;
    gap: 0.6rem;
    padding: 0.5rem 0 1.25rem 0;
}

.sl-logo-square {
    width: 34px;
    height: 34px;
    background: var(--sl-accent);
    color: #fff;
    border-radius: 10px;
    display: flex;
    align-items: center;
    justify-content: center;
    font-weight: 700;
    font-size: 1.1rem;
    flex-shrink: 0;
}

.sl-logo-text { line-height: 1.15; }
.sl-logo-title { font-weight: 700; font-size: 1.05rem; color: var(--sl-text); }
.sl-logo-subtitle { font-size: 0.78rem; color: var(--sl-text-muted); }

section[data-testid="stSidebar"] div[role="radiogroup"] {
    gap: 0.2rem;
}

section[data-testid="stSidebar"] div[role="radiogroup"] label {
    padding: 0.55rem 0.9rem;
    border-radius: 10px;
    margin-bottom: 0.1rem;
    font-weight: 500;
    transition: background-color 0.15s ease, color 0.15s ease;
}

section[data-testid="stSidebar"] div[role="radiogroup"] label > div:first-of-type {
    display: none;
}

section[data-testid="stSidebar"] div[role="radiogroup"] label:hover {
    background-color: #F1EEFC;
}

section[data-testid="stSidebar"] div[role="radiogroup"] label:has(input:checked) {
    background-color: var(--sl-accent);
}

section[data-testid="stSidebar"] div[role="radiogroup"] label:has(input:checked) p {
    color: #ffffff !important;
    font-weight: 600;
}

.sl-status-row {
    display: flex;
    align-items: center;
    gap: 0.5rem;
    font-size: 0.85rem;
    color: var(--sl-text-muted);
    margin-bottom: 0.4rem;
}

.sl-dot {
    width: 8px;
    height: 8px;
    border-radius: 50%;
    flex-shrink: 0;
}

.sl-dot-green { background-color: #22C55E; }
.sl-dot-red { background-color: #EF4444; }
.sl-dot-gray { background-color: #B9BFC9; }

div[data-testid="stVerticalBlockBorderWrapper"],
div[data-testid="stExpander"],
div[data-testid="stMetric"],
div[data-testid="stDataFrame"],
div[data-testid="stDataEditor"] {
    background-color: var(--sl-surface) !important;
    border-radius: 14px !important;
    border: 1px solid var(--sl-border) !important;
    box-shadow: none !important;
}

div[data-testid="stVerticalBlockBorderWrapper"] {
    padding: 0.5rem 0.75rem;
}

div.stButton > button, div.stDownloadButton > button {
    border-radius: 999px;
    font-weight: 600;
    border: 1px solid var(--sl-border);
    padding: 0.5rem 1.25rem;
    background-color: #ffffff;
    color: var(--sl-text);
}

div.stButton > button[kind="primary"], div.stDownloadButton > button[kind="primary"] {
    background-color: var(--sl-accent);
    border: none;
    color: #ffffff;
    box-shadow: none;
}

div.stButton > button[kind="primary"]:hover, div.stDownloadButton > button[kind="primary"]:hover {
    background-color: var(--sl-accent-dark);
    color: #ffffff;
}

div[data-testid="stMetric"] {
    padding: 1rem 1.25rem;
}

div[data-testid="stDataFrame"],
div[data-testid="stDataEditor"] {
    overflow: hidden;
}

div[data-testid="stTextInput"] input,
div[data-testid="stNumberInput"] input,
textarea,
div[data-testid="stSelectbox"] div[data-baseweb="select"] > div {
    border-radius: 10px !important;
    border-color: var(--sl-border) !important;
    background-color: #ffffff !important;
}

/* Ukrycie dodatkowych obramowań w tabelach Streamlit */
[data-testid="stDataFrame"] [data-testid="stTable"] {
    border: none !important;
}

/* Dopasowanie nagłówków tabeli */
[data-testid="stDataFrame"] th, [data-testid="stDataEditor"] th {
    background-color: #F9FAFB !important;
    color: var(--sl-text-muted) !important;
    font-weight: 600 !important;
    text-transform: uppercase;
    font-size: 0.75rem;
    letter-spacing: 0.025em;
}

.sl-pill {
    display: inline-block;
    padding: 2px 10px;
    border-radius: 999px;
    font-size: 0.82rem;
    font-weight: 600;
    white-space: nowrap;
}
</style>
"""
st.markdown(SELLTIC_CSS, unsafe_allow_html=True)

COLUMNS = ["place_id", "Nazwa", "Telefon", "Strona WWW", "Adres", "Ocena", "Liczba opinii", "Status", "Branża", "Miasto", "Data dodania", "Website Status", "Lead Score"]
EXPORT_COLUMNS = ["Nazwa", "Telefon", "Strona WWW", "Adres", "Ocena", "Liczba opinii", "Status", "Branża", "Miasto", "Data dodania", "Website Status", "Lead Score"]
MASTER_FILE = "baza_leadow.csv"
HISTORY_FILE = "historia_zapytan.json"
WEIGHTS_FILE = "scoring_weights.json"
CONFIG_FILE = "config.json"

# ── Hasło dostępu (Cloud Run: ustaw zmienną środowiskową APP_PASSWORD) ────────
# Wyjątek: jeśli aplikacja jest osadzona jako zakładka w CRM (iframe/reverse proxy),
# CRM przekazuje ?token=... zgodny z EMBED_TOKEN i logowanie jest pomijane,
# ponieważ użytkownik jest już uwierzytelniony po stronie CRM.

def check_password() -> bool:
    """Prosta bramka hasłem. Hasło bierze ze zmiennej środowiskowej APP_PASSWORD."""
    correct_password = os.environ.get("APP_PASSWORD", "")
    embed_token = os.environ.get("EMBED_TOKEN", "")

    if st.session_state.get("password_correct", False):
        return True

    url_token = st.query_params.get("token", "")
    if embed_token and url_token == embed_token:
        st.session_state["password_correct"] = True
        return True

    if not correct_password:
        # Brak ustawionego hasła w środowisku - nie blokuj (np. lokalny dev),
        # ale wypisz ostrzeżenie żeby nie zapomnieć go ustawić na produkcji.
        st.sidebar.warning("⚠️ APP_PASSWORD nie jest ustawione — aplikacja działa bez hasła!")
        return True

    def password_entered():
        if st.session_state.get("password_input") == correct_password:
            st.session_state["password_correct"] = True
            del st.session_state["password_input"]
        else:
            st.session_state["password_correct"] = False

    st.title("🔒 Selltic Scraper")
    st.text_input("Hasło dostępu", type="password", on_change=password_entered, key="password_input")
    if st.session_state.get("password_correct") is False:
        st.error("Niepoprawne hasło")
    return False


if not check_password():
    st.stop()


# ── Trwałość danych w Google Cloud Storage (opcjonalne, ale zalecane na Cloud Run) ──
# Cloud Run nie ma trwałego dysku - bez tego baza_leadow.csv i historia_zapytan.json
# znikną przy każdym restarcie/redeployu/skalowaniu kontenera.
# Ustaw zmienną środowiskową GCS_BUCKET, aby włączyć automatyczny zapis/odczyt z bucketa.

GCS_BUCKET = os.environ.get("GCS_BUCKET", "")
_gcs_synced = False


def _gcs_client():
    from google.cloud import storage
    return storage.Client()


def gcs_pull_all():
    """Pobiera baza_leadow.csv i historia_zapytan.json z bucketa, jeśli istnieją."""
    global _gcs_synced
    if not GCS_BUCKET or _gcs_synced:
        return
    try:
        client = _gcs_client()
        bucket = client.bucket(GCS_BUCKET)
        for fname in (MASTER_FILE, HISTORY_FILE, WEIGHTS_FILE):
            blob = bucket.blob(fname)
            if blob.exists():
                blob.download_to_filename(fname)
    except Exception as e:
        st.sidebar.error(f"Błąd pobierania danych z GCS: {e}")
    _gcs_synced = True


def gcs_push(local_filename: str):
    """Wysyła pojedynczy plik do bucketa po zapisie lokalnym."""
    if not GCS_BUCKET:
        return
    try:
        client = _gcs_client()
        bucket = client.bucket(GCS_BUCKET)
        blob = bucket.blob(local_filename)
        if os.path.exists(local_filename):
            blob.upload_from_filename(local_filename)
    except Exception as e:
        st.sidebar.error(f"Błąd zapisu danych do GCS: {e}")


gcs_pull_all()


def load_config() -> dict:
    """Wczytuje config.json (np. klucz API) z bucketa GCS, a jeśli GCS nie jest skonfigurowany - z pliku lokalnego."""
    if GCS_BUCKET:
        try:
            client = _gcs_client()
            blob = client.bucket(GCS_BUCKET).blob(CONFIG_FILE)
            if blob.exists():
                return json.loads(blob.download_as_text())
        except Exception as e:
            st.sidebar.error(f"Błąd odczytu config.json z GCS: {e}")
        return {}
    if os.path.exists(CONFIG_FILE):
        try:
            with open(CONFIG_FILE, "r", encoding="utf-8") as f:
                return json.load(f)
        except Exception:
            pass
    return {}


def save_config(data: dict):
    """Zapisuje config.json (np. klucz API) do bucketa GCS, a jeśli GCS nie jest skonfigurowany - do pliku lokalnego."""
    if GCS_BUCKET:
        try:
            client = _gcs_client()
            blob = client.bucket(GCS_BUCKET).blob(CONFIG_FILE)
            blob.upload_from_string(json.dumps(data, ensure_ascii=False, indent=2), content_type="application/json")
        except Exception as e:
            st.sidebar.error(f"Błąd zapisu config.json do GCS: {e}")
        return
    with open(CONFIG_FILE, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=2)


# ── Scoring stron WWW ──────────────────────────────────────────────────────────
# W pełni edytowalny z poziomu UI (zakładka Ustawienia → Konfiguracja scoringu):
# trzy stany strony WWW + osobny bonus za brak mobilności, oraz otwarte listy
# reguł punktowych dla liczby opinii i oceny Google.

DEFAULT_WEIGHTS = {
    "brak_strony": 40,
    "strona_nie_dziala": 30,
    "strona_dziala": 0,
    "niemobilna_bonus": 10,
    "reguly_opinii": [
        {"min_count": 1, "points": 5},
        {"min_count": 15, "points": 12},
        {"min_count": 50, "points": 20},
    ],
    "reguly_oceny": [
        {"min_rating": 3.0, "points": 5},
        {"min_rating": 4.0, "points": 10},
        {"min_rating": 4.5, "points": 15},
    ],
}


def load_weights() -> dict:
    """Wczytuje wagi scoringu z pliku, a jeśli brak - zapisuje i zwraca domyślne (seed)."""
    if os.path.exists(WEIGHTS_FILE):
        try:
            with open(WEIGHTS_FILE, "r", encoding="utf-8") as f:
                weights = json.load(f)
            merged = dict(DEFAULT_WEIGHTS)
            merged.update(weights)
            return merged
        except Exception:
            pass
    save_weights(DEFAULT_WEIGHTS)
    return dict(DEFAULT_WEIGHTS)


def save_weights(weights: dict):
    with open(WEIGHTS_FILE, "w", encoding="utf-8") as f:
        json.dump(weights, f, ensure_ascii=False, indent=2)
    gcs_push(WEIGHTS_FILE)


def _match_rule(rules: list[dict], value: float, threshold_key: str) -> tuple[int, dict | None]:
    """Znajduje regułę z najwyższym progiem <= value. Zwraca (punkty, reguła) albo (0, None)."""
    best = None
    for rule in rules:
        try:
            threshold = float(rule[threshold_key])
        except (KeyError, ValueError, TypeError):
            continue
        if threshold <= value and (best is None or threshold > float(best[threshold_key])):
            best = rule
    if best is None:
        return 0, None
    return int(best["points"]), best


def score_website(url: str, rating, review_count, weights: dict = None) -> tuple[int, str, dict]:
    """
    Zwraca (lead_score, website_status, breakdown).
    website_status: 'brak' | 'nie_dziala' | 'dziala'
    breakdown: {klucz: {"punkty": int, "opis": str}}
    """
    w = weights if weights is not None else load_weights()
    breakdown = {}

    if not url:
        website_status = "brak"
        score_website_pts = int(w["brak_strony"])
        breakdown["stan_strony"] = {"punkty": score_website_pts, "opis": "Brak strony/domeny"}
    else:
        # Uzupełnij brakujący schemat (np. "przyklad.pl" -> "https://przyklad.pl"),
        # inaczej requests rzuca MissingSchema i strona zawsze wygląda na "nie_dziala".
        if "://" not in url:
            url = "https://" + url
        try:
            resp = requests.get(url, timeout=5, headers={"User-Agent": "Mozilla/5.0"}, allow_redirects=True)
            if not resp.ok:
                website_status = "nie_dziala"
                score_website_pts = int(w["strona_nie_dziala"])
                breakdown["stan_strony"] = {"punkty": score_website_pts, "opis": "Jest domena, ale nie działa"}
            else:
                website_status = "dziala"
                score_website_pts = int(w["strona_dziala"])
                breakdown["stan_strony"] = {"punkty": score_website_pts, "opis": "Jest strona i działa"}
                html = resp.text[:20000]
                if 'name="viewport"' not in html.lower():
                    bonus = int(w["niemobilna_bonus"])
                    score_website_pts += bonus
                    breakdown["niemobilna"] = {"punkty": bonus, "opis": "Strona niemobilna (brak meta viewport)"}
        except Exception:
            website_status = "nie_dziala"
            score_website_pts = int(w["strona_nie_dziala"])
            breakdown["stan_strony"] = {"punkty": score_website_pts, "opis": "Jest domena, ale nie działa (błąd połączenia)"}

    try:
        rc = int(review_count) if review_count not in (None, "") else 0
    except (ValueError, TypeError):
        rc = 0
    try:
        rt = float(rating) if rating not in (None, "") else 0
    except (ValueError, TypeError):
        rt = 0

    opinie_pts, opinie_rule = _match_rule(w.get("reguly_opinii", []), rc, "min_count")
    if opinie_rule is not None:
        breakdown["opinie"] = {"punkty": opinie_pts, "opis": f"≥ {opinie_rule['min_count']} opinii"}

    ocena_pts, ocena_rule = _match_rule(w.get("reguly_oceny", []), rt, "min_rating")
    if ocena_rule is not None:
        breakdown["ocena"] = {"punkty": ocena_pts, "opis": f"≥ {ocena_rule['min_rating']} oceny"}

    total = score_website_pts + opinie_pts + ocena_pts
    return total, website_status, breakdown


def format_breakdown(breakdown: dict) -> list[str]:
    return [f"+{item['punkty']} pkt: {item['opis']}" for item in breakdown.values()]


# ── Integracja z CRM (selltic-crm) ──────────────────────────────────────────────
# Wymagane zmienne środowiskowe: CRM_API_BASE_URL, SCRAPER_IMPORT_KEY
# (ta sama wartość SCRAPER_IMPORT_KEY musi być ustawiona też w Vercel po stronie CRM)

CRM_API_BASE_URL = os.environ.get("CRM_API_BASE_URL", "").rstrip("/")
SCRAPER_IMPORT_KEY = os.environ.get("SCRAPER_IMPORT_KEY", "")
CRM_ENABLED = bool(CRM_API_BASE_URL and SCRAPER_IMPORT_KEY)


def crm_test_connection() -> tuple[bool, str]:
    """
    Sprawdza połączenie z CRM wywołując existing-ids z pustą listą place_id. Zwraca (ok, opis).
    UWAGA: Ten test sprawdza tylko łączność i klucz API, nie waliduje schematu importu.
    """
    if not CRM_ENABLED:
        return False, "CRM nieskonfigurowany (brak CRM_API_BASE_URL / SCRAPER_IMPORT_KEY)."
    try:
        resp = requests.get(
            f"{CRM_API_BASE_URL}/api/prospecting/existing-ids",
            params={"place_ids": ""},
            headers={"X-API-Key": SCRAPER_IMPORT_KEY},
            timeout=10,
        )
        if resp.ok:
            return True, f"Połączono z CRM ({CRM_API_BASE_URL})."
        return False, f"CRM odpowiedział błędem: {resp.status_code}"
    except Exception as e:
        return False, f"Nie udało się połączyć z CRM: {e}"


def crm_check_existing(place_ids: list[str]) -> set[str]:
    """Pyta CRM które z podanych place_id już istnieją. Zwraca pusty set przy błędzie/braku konfiguracji."""
    if not CRM_ENABLED or not place_ids:
        return set()
    try:
        resp = requests.get(
            f"{CRM_API_BASE_URL}/api/prospecting/existing-ids",
            params={"place_ids": ",".join(place_ids)},
            headers={"X-API-Key": SCRAPER_IMPORT_KEY},
            timeout=10,
        )
        if resp.ok:
            return set(resp.json().get("existing", []))
    except Exception as e:
        st.sidebar.warning(f"⚠️ Nie udało się sprawdzić duplikatów w CRM: {e}")
    return set()


def _prepare_crm_payload(rows: list[dict]) -> list[dict]:
    """Ujednolica format danych wysyłanych do CRM."""
    payload = []
    for r in rows:
        place_id = r.get("place_id", "")
        name = r.get("Nazwa", "")
        city = r.get("Miasto", "")
        industry = r.get("Branża", "")

        website = r.get("Strona WWW", "")
        website = None if website in ("", "BRAK") else website

        try:
            rating = float(r["Ocena"]) if r.get("Ocena") not in (None, "") else None
        except (ValueError, TypeError):
            rating = None
        try:
            review_count = int(float(r["Liczba opinii"])) if r.get("Liczba opinii") not in (None, "") else None
        except (ValueError, TypeError):
            review_count = None
        try:
            score_val = int(float(r.get("Lead Score", 0))) if r.get("Lead Score") not in (None, "") else None
        except (ValueError, TypeError):
            score_val = None

        if score_val is None:
            priority_label = None
        elif score_val >= 70:
            priority_label = "wysoki"
        elif score_val >= 40:
            priority_label = "średni"
        else:
            priority_label = "niski"

        if place_id:
            google_maps_url = f"https://maps.google.com/?cid={place_id}"
        else:
            google_maps_url = f"https://www.google.com/maps/search/?api=1&query={quote(f'{name} {city}')}"

        payload.append({
            "place_id": place_id,
            "name": name,
            "phone": r.get("Telefon") or None,
            "website": website,
            "address": r.get("Adres", ""),
            "rating": rating,
            "review_count": review_count,
            "business_status": r.get("Status", ""),
            "industry": industry,
            "category": industry,
            "city": city,
            "google_maps_url": google_maps_url,
            "lead_score": score_val,
            "priority_score": score_val,
            "priority_label": priority_label,
            "website_status": r.get("Website Status") or None,
            "score_reasons": [],
        })
    return payload


def crm_import_batch(rows: list[dict]) -> dict | None:
    """Wysyła paczkę leadów do CRM. Zwraca odpowiedź albo None przy błędzie/braku konfiguracji."""
    if not CRM_ENABLED or not rows:
        return None
    payload = _prepare_crm_payload(rows)
    try:
        resp = requests.post(
            f"{CRM_API_BASE_URL}/api/prospecting/import",
            json=payload,
            headers={"X-API-Key": SCRAPER_IMPORT_KEY},
            timeout=30,
        )
        if resp.ok:
            return resp.json()
        st.error(f"❌ CRM odpowiedział błędem: {resp.status_code}")
        st.sidebar.error(f"❌ CRM błąd: {resp.status_code}")
    except Exception as e:
        st.error(f"❌ Nie udało się wysłać leadów do CRM: {e}")
        st.sidebar.error(f"❌ CRM błąd połączenia")
    return None


# ── Wysyłka wybranych leadów z zakładki "Baza leadów" do CRM (prospecting/import) ──

def build_prospecting_payload(rows: list[dict]) -> list[dict]:
    """Mapuje zaznaczone wiersze z tabeli leadów na format /api/prospecting/import."""
    return _prepare_crm_payload(rows)


def build_results_table(rows: list[dict], result_data: dict) -> pd.DataFrame:
    """Buduje tabelę wyników (Nazwa | Status | Info) na podstawie odpowiedzi CRM."""
    status_labels = {
        "created": "✅ Dodano",
        "updated": "🔄 Zaktualizowano",
        "error": "❌ Błąd",
    }
    items = result_data.get("results") if isinstance(result_data, dict) else None

    table_rows = []
    if isinstance(items, list) and len(items) == len(rows):
        for row, item in zip(rows, items):
            status_key = str(item.get("status", "")).lower()
            status_label = status_labels.get(status_key, item.get("status", "?"))
            info = item.get("error") or item.get("message") or ""
            table_rows.append({"Nazwa": row.get("Nazwa", ""), "Status": status_label, "Info": info})
    else:
        added = result_data.get("added") if isinstance(result_data, dict) else None
        updated = result_data.get("updated") if isinstance(result_data, dict) else None
        info = f"Dodano: {added if added is not None else '?'}, zaktualizowano: {updated if updated is not None else '?'}"
        for row in rows:
            table_rows.append({"Nazwa": row.get("Nazwa", ""), "Status": "✅ Wysłano", "Info": info})
    return pd.DataFrame(table_rows)


def send_prospecting_selection(rows: list[dict]) -> tuple[bool, str, pd.DataFrame | None]:
    """Wysyła ręcznie zaznaczone leady do CRM (/api/prospecting/import). Zwraca (ok, komunikat, tabela wynikowa)."""
    payload = build_prospecting_payload(rows)
    try:
        resp = requests.post(
            f"{CRM_API_BASE_URL}/api/prospecting/import",
            json=payload,
            headers={"X-API-Key": SCRAPER_IMPORT_KEY},
            timeout=30,
        )
    except Exception as e:
        return False, f"❌ Nie udało się połączyć z CRM (Timeout/DNS): {e}", None
    if resp.ok:
        try:
            result_data = resp.json()
        except Exception:
            result_data = {}
        return True, "📤 Wysłano do CRM.", build_results_table(rows, result_data)
    return False, f"❌ CRM odpowiedział błędem {resp.status_code}: {resp.text}", None


# ── Master DB helpers ─────────────────────────────────────────────────────────

def load_master() -> pd.DataFrame:
    if os.path.exists(MASTER_FILE):
        df = pd.read_csv(MASTER_FILE, dtype=str).fillna("")
        for col in COLUMNS:
            if col not in df.columns:
                df[col] = ""
        return df[COLUMNS]
    return pd.DataFrame(columns=COLUMNS)


def save_master(df: pd.DataFrame):
    df.to_csv(MASTER_FILE, index=False, encoding="utf-8-sig")
    gcs_push(MASTER_FILE)


def append_to_master(rows: list[dict]) -> tuple[int, int]:
    """Append new rows to master, deduplicate by place_id. Returns (added, skipped)."""
    df = load_master()
    existing_ids = set(df["place_id"].tolist())
    new_rows, skipped = [], 0
    for row in rows:
        if row["place_id"] in existing_ids:
            skipped += 1
        else:
            existing_ids.add(row["place_id"])
            new_rows.append(row)
    if new_rows:
        df = pd.concat([df, pd.DataFrame(new_rows, columns=COLUMNS)], ignore_index=True)
        save_master(df)
    return len(new_rows), skipped


# ── History helpers ───────────────────────────────────────────────────────────

def load_history() -> dict:
    if os.path.exists(HISTORY_FILE):
        with open(HISTORY_FILE, "r", encoding="utf-8") as f:
            return json.load(f)
    return {}


def save_query_to_history(branza: str, lok: str, leady: int):
    history = load_history()
    key = f"{branza.lower().strip()}|{lok.lower().strip()}"
    history[key] = {"data": datetime.now().strftime("%Y-%m-%d %H:%M"), "leady": leady}
    with open(HISTORY_FILE, "w", encoding="utf-8") as f:
        json.dump(history, f, ensure_ascii=False, indent=2)
    gcs_push(HISTORY_FILE)


def qkey(branza: str, lok: str) -> str:
    return f"{branza.lower().strip()}|{lok.lower().strip()}"


# ── API helpers ───────────────────────────────────────────────────────────────

def text_search(query, api_key, page_token=None):
    url = "https://maps.googleapis.com/maps/api/place/textsearch/json"
    params = {"query": query, "key": api_key, "language": "pl"}
    if page_token:
        params["pagetoken"] = page_token
    return requests.get(url, params=params, timeout=10).json()


def place_details(place_id, api_key):
    url = "https://maps.googleapis.com/maps/api/place/details/json"
    params = {
        "place_id": place_id,
        "fields": "name,formatted_phone_number,website,formatted_address,rating,user_ratings_total,business_status",
        "key": api_key,
        "language": "pl"
    }
    return requests.get(url, params=params, timeout=10).json().get("result", {})


# ── Excel export ──────────────────────────────────────────────────────────────

def to_excel(df: pd.DataFrame) -> BytesIO:
    wb = Workbook()
    ws = wb.active
    ws.title = "Leady"
    header_fill = PatternFill("solid", fgColor="1A3ADB")
    header_font = Font(bold=True, color="FFFFFF", name="Arial", size=10)
    border = Border(
        left=Side(style="thin", color="DDDDDD"),
        right=Side(style="thin", color="DDDDDD"),
        bottom=Side(style="thin", color="DDDDDD"),
    )
    cols = list(df.columns)
    for ci, col in enumerate(cols, 1):
        cell = ws.cell(row=1, column=ci, value=col)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
    for ri, row in enumerate(df.itertuples(index=False), 2):
        for ci, val in enumerate(row, 1):
            cell = ws.cell(row=ri, column=ci, value=val)
            cell.font = Font(name="Arial", size=10)
            cell.border = border
            if ri % 2 == 0:
                cell.fill = PatternFill("solid", fgColor="F0F4FF")
    col_widths = {"Nazwa": 30, "Telefon": 18, "Strona WWW": 35, "Adres": 40,
                  "Ocena": 10, "Liczba opinii": 14, "Status": 18, "Branża": 16,
                  "Miasto": 16, "Data dodania": 18}
    for ci, col in enumerate(cols, 1):
        ws.column_dimensions[get_column_letter(ci)].width = col_widths.get(col, 15)
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


# ── UI display helpers (badge/pill rendering) ─────────────────────────────────

def status_dot(color: str) -> str:
    return f'<span class="sl-dot sl-dot-{color}"></span>'


def status_line(color: str, text: str) -> str:
    return f'<div class="sl-status-row">{status_dot(color)}{text}</div>'


def pill_html(text, color: str, bg: str) -> str:
    return f'<span class="sl-pill" style="color:{color};background:{bg};">{text}</span>'


def score_pill(score) -> str:
    try:
        s = int(float(score))
    except (ValueError, TypeError):
        s = 0
    if s >= 70:
        return pill_html(s, "#1B8A5A", "#E6F7EF")
    if s >= 35:
        return pill_html(s, "#B7791F", "#FEF3D7")
    return pill_html(s, "#5B6472", "#F1F2F4")


def website_badge(website: str) -> str:
    if not website or website == "BRAK":
        return pill_html("Brak strony", "#1B8A5A", "#E6F7EF")
    domain = str(website).replace("https://", "").replace("http://", "").split("/")[0]
    return pill_html(domain, "#5B6472", "#F1F2F4")


def maps_url(place_id: str, name: str, city: str) -> str:
    if place_id:
        return f"https://maps.google.com/?cid={place_id}"
    return f"https://www.google.com/maps/search/?api=1&query={quote(f'{name} {city}')}"


def render_leads_rows(rows: list[dict], key_prefix: str):
    """Renderuje tabelę leadów jako wiersze z checkboxami + kolorowymi odznakami (Score/Strona WWW)."""
    widths = [0.4, 3, 1.5, 2, 1.8, 1, 0.8]
    header_cols = st.columns(widths)
    for c, h in zip(header_cols, ["", "Firma", "Telefon", "Strona WWW", "Ocena / opinie", "Score", "Mapy"]):
        c.markdown(f"**{h}**")
    st.markdown('<hr style="margin:0.25rem 0;border:none;border-top:1px solid #ECEEF3;">', unsafe_allow_html=True)
    for row in rows:
        cols = st.columns(widths)
        cols[0].checkbox("Wybierz", key=f"{key_prefix}_{row.get('place_id','')}", label_visibility="collapsed")
        cols[1].markdown(
            f"**{row.get('Nazwa','')}**<br><span style='color:#767E8C;font-size:0.82rem;'>{row.get('Branża','')}</span>",
            unsafe_allow_html=True,
        )
        cols[2].markdown(row.get("Telefon") or "—")
        cols[3].markdown(website_badge(row.get("Strona WWW", "")), unsafe_allow_html=True)
        rating = row.get("Ocena") or "—"
        reviews = row.get("Liczba opinii") or 0
        cols[4].markdown(f"⭐ {rating} ({reviews})")
        cols[5].markdown(score_pill(row.get("Lead Score", 0)), unsafe_allow_html=True)
        link = maps_url(row.get("place_id", ""), row.get("Nazwa", ""), row.get("Miasto", ""))
        cols[6].markdown(f"[📍]({link})")
        st.markdown('<hr style="margin:0.25rem 0;border:none;border-top:1px solid #ECEEF3;">', unsafe_allow_html=True)


# ══════════════════════════════════════════════════════════════════════════════
# UI
# ══════════════════════════════════════════════════════════════════════════════

env_api_key = os.environ.get("GOOGLE_PLACES_API_KEY", "")
app_config = load_config()
# Single source of truth: session_state["manual_api_key"] is authoritative once set.
# Self-heal only if session_state is missing/empty but config.json (GCS/local) now has a key -
# this covers a fresh session/container that hasn't picked up a key saved from another session.
if "manual_api_key" not in st.session_state or (
    not st.session_state.get("manual_api_key") and app_config.get("google_places_api_key")
):
    st.session_state["manual_api_key"] = app_config.get("google_places_api_key", "")
api_key = env_api_key if env_api_key else st.session_state.get("manual_api_key", "")

_api_key_source = "env" if env_api_key else ("session_state" if st.session_state.get("manual_api_key") else "config/empty")
print(f"[api_key debug] source={_api_key_source} empty={not bool(api_key)}")

PAGES = [
    "🚀 Scraper",
    "📦 Baza leadów",
    "📋 Historia zapytań",
    "⚙️ Ustawienia",
]

with st.sidebar:
    st.markdown(
        '<div class="sl-logo-row">'
        '<div class="sl-logo-square">S</div>'
        '<div class="sl-logo-text">'
        '<div class="sl-logo-title">Selltic</div>'
        '<div class="sl-logo-subtitle">Scraper</div>'
        '</div></div>',
        unsafe_allow_html=True,
    )
    page = st.radio("Nawigacja", PAGES, label_visibility="collapsed", key="nav_page")

    st.markdown("<div style='height:1rem;'></div>", unsafe_allow_html=True)

    if api_key:
        st.markdown(status_line("green", "Klucz API zapisany"), unsafe_allow_html=True)
    else:
        st.markdown(status_line("red", "Brak klucza API"), unsafe_allow_html=True)

    if CRM_ENABLED:
        st.markdown(status_line("green", "CRM połączony"), unsafe_allow_html=True)
    else:
        st.markdown(status_line("gray", "CRM nieskonfigurowany"), unsafe_allow_html=True)

    st.caption(f"v · {os.environ.get('APP_VERSION') or 'local'}")

st.title("Selltic – Google Maps Scraper")


# ── Strona: Scraper ───────────────────────────────────────────────────────────
if page == "🚀 Scraper":
    history = load_history()
    master_ids = set(load_master()["place_id"].tolist())

    if "last_session_rows" not in st.session_state:
        st.session_state["last_session_rows"] = []

    with st.container(border=True):
        col1, col2 = st.columns(2)
        with col1:
            branze_input = st.text_area(
                "Branża", key="input_branza",
                placeholder="hydraulik\nusługi hydrauliczne\ninstalator wod-kan", height=110,
            )
        with col2:
            lokalizacje_input = st.text_area(
                "Miasto", key="input_miasto",
                placeholder="Wrocław Krzyki\nWrocław Fabryczna\nWrocław Śródmieście", height=110,
            )

        col3, col4 = st.columns([1, 2])
        with col3:
            limit_firm = st.number_input("Limit firm", min_value=20, max_value=180, value=60, step=20)
            st.caption("Google Places zwraca ~20 firm na stronę")
        with col4:
            powtorz = st.checkbox("Wykonaj ponownie już zrobione zapytania", value=False)

        max_strony = min(3, math.ceil(limit_firm / 20))

        run_clicked = st.button("🚀 Szukaj", type="primary", use_container_width=True)

    auto_run = st.session_state.pop("auto_run_scraper", False)
    should_run = run_clicked or auto_run

    # Ostatnie wyszukiwania — klikalne chipy powtarzające zapytanie
    if history:
        st.markdown("**Ostatnie wyszukiwania — kliknij żeby powtórzyć**")
        recent = sorted(history.items(), key=lambda x: x[1]["data"], reverse=True)[:8]
        chip_cols = st.columns(4)
        for i, (k, v) in enumerate(recent):
            branza_h, miasto_h = k.split("|", 1)
            label = f"{branza_h.title()} · {miasto_h.title()} — {v['leady']} firm — {v['data']}"
            if chip_cols[i % 4].button(label, key=f"chip_{i}", use_container_width=True):
                st.session_state["input_branza"] = branza_h
                st.session_state["input_miasto"] = miasto_h
                st.session_state["auto_run_scraper"] = True
                st.rerun()

    # Query preview
    if branze_input.strip() and lokalizacje_input.strip():
        branze_prev = [b.strip() for b in branze_input.strip().splitlines() if b.strip()]
        lok_prev = [l.strip() for l in lokalizacje_input.strip().splitlines() if l.strip()]
        nowe, zrobione = [], []
        for b in branze_prev:
            for l in lok_prev:
                k = qkey(b, l)
                if k in history:
                    zrobione.append(f"{b} · {l} — {history[k]['data']}, {history[k]['leady']} leadów")
                else:
                    nowe.append(f"{b} · {l}")
        if zrobione:
            with st.expander(f"⚠️ {len(zrobione)} zapytań już wykonanych – zostaną pominięte"):
                for z in zrobione:
                    st.markdown(f"- ~~{z}~~")
        if nowe:
            with st.expander(f"✅ {len(nowe)} nowych zapytań", expanded=True):
                for n in nowe:
                    st.markdown(f"- {n}")

    if should_run:
        if not api_key:
            st.error("Wklej klucz API w panelu bocznym.")
        elif not branze_input.strip() or not lokalizacje_input.strip():
            st.error("Uzupełnij branże i lokalizacje.")
        else:
            branze = [b.strip() for b in branze_input.strip().splitlines() if b.strip()]
            lokalizacje = [l.strip() for l in lokalizacje_input.strip().splitlines() if l.strip()]
            pairs = [(b, l) for b in branze for l in lokalizacje
                     if qkey(b, l) not in history or powtorz]

            if not pairs:
                st.warning("Wszystkie zapytania już wykonane. Zaznacz 'Wykonaj ponownie' żeby powtórzyć.")
                st.stop()

            st.info(f"💾 Wyniki trafią do **{MASTER_FILE}** · do wykonania: **{len(pairs)} zapytań**")
            progress_bar = st.progress(0)
            status_text = st.empty()
            m1, m2, m3, m4 = st.columns(4)
            met_dodane = m1.empty()
            met_duplikaty = m2.empty()
            met_query = m3.empty()
            met_pominiete = m4.empty()
            table_placeholder = st.empty()

            session_rows = []
            seen_this_session = set()
            pominiete = 0
            duplikaty_baza = 0

            for pair_nr, (branza, lok) in enumerate(pairs, 1):
                query = f"{branza} {lok}"
                next_token = None
                leady_dla_query = 0

                for page_nr in range(max_strony):
                    status_text.info(f"⏳ {pair_nr}/{len(pairs)}: **{query}** – strona {page_nr + 1}")
                    if page_nr > 0 and next_token:
                        for i in range(2, 0, -1):
                            status_text.info(f"⏳ **{query}** – czekam {i}s...")
                            time.sleep(1)

                    try:
                        data = text_search(query, api_key, next_token if page_nr > 0 else None)
                    except Exception as e:
                        status_text.error(f"❌ Błąd: {e}")
                        break

                    api_status = data.get("status")
                    if api_status == "REQUEST_DENIED":
                        status_text.error("❌ REQUEST_DENIED – sprawdź klucz API i rozliczenia.")
                        st.stop()
                    elif api_status in ("ZERO_RESULTS",):
                        break
                    elif api_status != "OK":
                        status_text.warning(f"⚠️ {api_status}: {data.get('error_message','')}")
                        break

                    results = data.get("results", [])

                    # Dedup względem CRM (jedno zapytanie na całą stronę wyników)
                    crm_existing_ids = crm_check_existing([p["place_id"] for p in results])

                    for i, place in enumerate(results):
                        pid = place["place_id"]
                        status_text.info(f"⏳ **{query}** – {i+1}/{len(results)}: *{place.get('name','')}*")

                        # Dedup within session
                        if pid in seen_this_session:
                            pominiete += 1
                            met_pominiete.metric("⏭️ Duplikaty sesji", pominiete)
                            continue
                        seen_this_session.add(pid)

                        # Dedup against master DB
                        if pid in master_ids:
                            duplikaty_baza += 1
                            met_duplikaty.metric("🗄️ Już w bazie", duplikaty_baza)
                            continue

                        # Dedup against CRM (oszczędza wywołania place_details/Places API)
                        if pid in crm_existing_ids:
                            duplikaty_baza += 1
                            met_duplikaty.metric("🗄️ Już w bazie", duplikaty_baza)
                            continue

                        try:
                            details = place_details(pid, api_key)
                        except Exception:
                            continue
                        time.sleep(0.05)

                        strona_www = details.get("website", "")

                        ocena = details.get("rating", "")
                        liczba_opinii = details.get("user_ratings_total", "")
                        lead_score, website_status, _breakdown = score_website(strona_www, ocena, liczba_opinii)

                        row = {
                            "place_id": pid,
                            "Nazwa": details.get("name", ""),
                            "Telefon": details.get("formatted_phone_number", ""),
                            "Strona WWW": strona_www if strona_www else "BRAK",
                            "Adres": details.get("formatted_address", ""),
                            "Ocena": ocena,
                            "Liczba opinii": liczba_opinii,
                            "Status": details.get("business_status", ""),
                            "Branża": branza,
                            "Miasto": lok,
                            "Data dodania": datetime.now().strftime("%Y-%m-%d %H:%M"),
                            "Website Status": website_status,
                            "Lead Score": lead_score,
                        }

                        session_rows.append(row)
                        master_ids.add(pid)
                        leady_dla_query += 1

                        # Append to master immediately
                        df_new = pd.DataFrame([row], columns=COLUMNS)
                        if os.path.exists(MASTER_FILE):
                            df_new.to_csv(MASTER_FILE, mode="a", header=False, index=False, encoding="utf-8-sig")
                        else:
                            df_new.to_csv(MASTER_FILE, index=False, encoding="utf-8-sig")
                        gcs_push(MASTER_FILE)

                        met_dodane.metric("✅ Dodano do bazy", len(session_rows))
                        met_duplikaty.metric("🗄️ Już w bazie", duplikaty_baza)
                        met_query.metric("🔍 Zapytanie", f"{pair_nr}/{len(pairs)}")
                        met_pominiete.metric("⏭️ Duplikaty sesji", pominiete)
                        table_placeholder.dataframe(
                            pd.DataFrame(session_rows)[EXPORT_COLUMNS],
                            use_container_width=True, height=350
                        )

                    next_token = data.get("next_page_token")
                    if not next_token:
                        break

                save_query_to_history(branza, lok, leady_dla_query)
                progress_bar.progress(pair_nr / len(pairs))

            progress_bar.progress(1.0)
            status_text.success(f"✅ Gotowe! Dodano **{len(session_rows)}** nowych leadów · pominięto **{duplikaty_baza}** już istniejących w bazie.")

            st.session_state["last_session_rows"] = session_rows

            if session_rows:
                # Automatyczny import do CRM
                if CRM_ENABLED:
                    with st.spinner("📤 Wysyłam leady do CRM..."):
                        result = crm_import_batch(session_rows)
                    if result is not None:
                        st.success(
                            f"📤 CRM: dodano **{result.get('added', '?')}**, "
                            f"zaktualizowano **{result.get('updated', '?')}**."
                        )
                        st.session_state.pop("crm_retry_rows", None)
                    else:
                        st.session_state["crm_retry_rows"] = session_rows
                        st.error("❌ Import do CRM nie powiódł się. Możesz spróbować ponownie poniżej.")

    # Przycisk ponowienia importu, jeśli ostatnia automatyczna wysyłka się nie powiodła
    if CRM_ENABLED and st.session_state.get("crm_retry_rows"):
        if st.button("🔁 Spróbuj ponownie wysłać do CRM", use_container_width=True):
            with st.spinner("📤 Wysyłam leady do CRM..."):
                result = crm_import_batch(st.session_state["crm_retry_rows"])
            if result is not None:
                st.success(f"📤 CRM: dodano **{result.get('added', '?')}**, zaktualizowano **{result.get('updated', '?')}**.")
                st.session_state.pop("crm_retry_rows", None)
            else:
                st.error("❌ Nadal nie udało się połączyć z CRM.")

    # Wyniki wyszukiwania (utrzymywane w session_state, żeby filtry/checkboxy działały bez ponownego scrapingu)
    if st.session_state["last_session_rows"]:
        st.divider()
        st.markdown("#### Wyniki wyszukiwania")

        fcol1, fcol2 = st.columns(2)
        with fcol1:
            filtr_priorytet = st.selectbox(
                "Priorytet",
                ["Wszystkie priorytety", "Wysoki (≥70)", "Średni (35-69)", "Niski (<35)"],
                key="res_filtr_priorytet",
            )
        with fcol2:
            filtr_strona_res = st.selectbox(
                "Strona WWW",
                ["Wszystkie strony", "Tylko bez strony", "Tylko ze stroną"],
                key="res_filtr_strona",
            )

        rows_view = list(st.session_state["last_session_rows"])

        def _score(r):
            try:
                return int(float(r.get("Lead Score") or 0))
            except (ValueError, TypeError):
                return 0

        if filtr_priorytet == "Wysoki (≥70)":
            rows_view = [r for r in rows_view if _score(r) >= 70]
        elif filtr_priorytet == "Średni (35-69)":
            rows_view = [r for r in rows_view if 35 <= _score(r) < 70]
        elif filtr_priorytet == "Niski (<35)":
            rows_view = [r for r in rows_view if _score(r) < 35]

        if filtr_strona_res == "Tylko bez strony":
            rows_view = [r for r in rows_view if not r.get("Strona WWW") or r.get("Strona WWW") == "BRAK"]
        elif filtr_strona_res == "Tylko ze stroną":
            rows_view = [r for r in rows_view if r.get("Strona WWW") and r.get("Strona WWW") != "BRAK"]

        with st.container(border=True):
            if rows_view:
                render_leads_rows(rows_view, key_prefix="res")
            else:
                st.caption("Brak wyników pasujących do filtrów.")

        selected_rows = [r for r in rows_view if st.session_state.get(f"res_{r.get('place_id','')}")]
        if selected_rows:
            if st.button(f"📤 Wyślij zaznaczone do CRM ({len(selected_rows)})", type="primary",
                         use_container_width=True, key="send_selected_scraper"):
                with st.spinner("📤 Wysyłam leady do CRM..."):
                    ok, msg, results_table = send_prospecting_selection(selected_rows)
                if ok:
                    st.success(msg)
                    if results_table is not None:
                        st.dataframe(results_table, use_container_width=True, hide_index=True)
                else:
                    st.error(msg)

        df_session_full = pd.DataFrame(st.session_state["last_session_rows"])[EXPORT_COLUMNS]
        excel_buf = to_excel(df_session_full)
        st.download_button(
            label="📥 Pobierz Excel (tylko ta sesja)",
            data=excel_buf,
            file_name=f"leady_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True,
        )


# ── Strona: Baza leadów ───────────────────────────────────────────────────────
elif page == "📦 Baza leadów":
    df_master = load_master()

    if df_master.empty:
        st.markdown(
            "<div style='text-align:center;padding:3.5rem 0;color:#767E8C;'>"
            "<div style='font-size:2.5rem;'>🔍</div>"
            "<div style='margin-top:0.75rem;'>Baza jest pusta — uruchom scraper żeby zebrać pierwsze leady</div>"
            "</div>",
            unsafe_allow_html=True,
        )
    else:
        col1, col2, col3, col4 = st.columns(4)
        col1.metric("Wszystkich leadów", len(df_master))
        col2.metric("Bez strony WWW", (df_master["Strona WWW"] == "BRAK").sum())
        col3.metric("Unikalne miasta", df_master["Miasto"].nunique())
        col4.metric("Unikalne branże", df_master["Branża"].nunique())

        st.divider()

        with st.container(border=True):
            # Filters
            fc1, fc2, fc3 = st.columns(3)
            with fc1:
                miasta = ["Wszystkie"] + sorted(df_master["Miasto"].unique().tolist())
                filtr_miasto = st.selectbox("Miasto", miasta)
            with fc2:
                branze_list = ["Wszystkie"] + sorted(df_master["Branża"].unique().tolist())
                filtr_branza = st.selectbox("Branża", branze_list)
            with fc3:
                filtr_strona = st.selectbox("Strona WWW", ["Wszystkie", "Tylko BEZ strony", "Tylko Z stroną"])

            df_filtered = df_master.copy()
            if filtr_miasto != "Wszystkie":
                df_filtered = df_filtered[df_filtered["Miasto"] == filtr_miasto]
            if filtr_branza != "Wszystkie":
                df_filtered = df_filtered[df_filtered["Branża"] == filtr_branza]
            if filtr_strona == "Tylko BEZ strony":
                df_filtered = df_filtered[df_filtered["Strona WWW"] == "BRAK"]
            elif filtr_strona == "Tylko Z stroną":
                df_filtered = df_filtered[df_filtered["Strona WWW"] != "BRAK"]

            st.markdown(f"**{len(df_filtered)} rekordów** po filtrach")

            df_filtered_reset = df_filtered.reset_index(drop=True)

            if CRM_ENABLED:
                df_display = df_filtered_reset[EXPORT_COLUMNS].copy()
                df_display.insert(0, "Wyślij", False)
                df_edited = st.data_editor(
                    df_display,
                    use_container_width=True, height=450, hide_index=True,
                    disabled=EXPORT_COLUMNS,
                    column_config={"Wyślij": st.column_config.CheckboxColumn("Wyślij", default=False)},
                    key="lead_table_editor",
                )

                if st.button("📤 Wyślij zaznaczone do CRM", type="primary", use_container_width=True):
                    selected_positions = df_edited.index[df_edited["Wyślij"] == True].tolist()
                    if not selected_positions:
                        st.warning("Zaznacz przynajmniej jeden wiersz w kolumnie 'Wyślij'.")
                    else:
                        selected_rows = df_filtered_reset.loc[selected_positions].to_dict("records")
                        with st.spinner(f"📤 Wysyłam {len(selected_rows)} leadów do CRM..."):
                            ok, msg, results_table = send_prospecting_selection(selected_rows)
                        if ok:
                            st.success(msg)
                            if results_table is not None:
                                st.dataframe(results_table, use_container_width=True, hide_index=True)
                        else:
                            st.error(msg)
            else:
                st.dataframe(df_filtered_reset[EXPORT_COLUMNS], use_container_width=True, height=450, hide_index=True)

        bc1, bc2, bc3 = st.columns(3)
        with bc1:
            excel_all = to_excel(df_filtered[EXPORT_COLUMNS])
            st.download_button("📥 Pobierz Excel", excel_all,
                file_name="baza_leadow.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True, type="primary")
        with bc2:
            csv_data = df_filtered[EXPORT_COLUMNS].to_csv(index=False, encoding="utf-8-sig").encode("utf-8-sig")
            st.download_button("📥 Pobierz CSV", csv_data,
                file_name="baza_leadow.csv", mime="text/csv", use_container_width=True)
        with bc3:
            if st.button("🗑️ Wyczyść całą bazę", use_container_width=True):
                os.remove(MASTER_FILE)
                if GCS_BUCKET:
                    try:
                        _gcs_client().bucket(GCS_BUCKET).blob(MASTER_FILE).delete()
                    except Exception:
                        pass
                st.rerun()


# ── Strona: Historia zapytań ──────────────────────────────────────────────────
elif page == "📋 Historia zapytań":
    history = load_history()

    if not history:
        st.markdown(
            "<div style='text-align:center;padding:3.5rem 0;color:#767E8C;'>"
            "<div style='font-size:2.5rem;'>📋</div>"
            "<div style='margin-top:0.75rem;'>Brak historii — uruchom scraper żeby zarejestrować pierwsze zapytania</div>"
            "</div>",
            unsafe_allow_html=True,
        )
    else:
        items = sorted(history.items(), key=lambda x: x[1]["data"], reverse=True)

        hc1, hc2 = st.columns(2)
        hc1.metric("Wykonanych zapytań", len(items))
        hc2.metric("Leadów łącznie z historii", sum(v["leady"] for _, v in items))

        st.divider()

        for i, (k, v) in enumerate(items):
            branza_h, miasto_h = k.split("|", 1)
            with st.container(border=True):
                hcol1, hcol2 = st.columns([4, 1])
                with hcol1:
                    st.markdown(f"**{branza_h.title()} · {miasto_h.title()}**")
                    st.caption(f"{v['data']} · {v['leady']} firm znalezionych")
                with hcol2:
                    if st.button("🔄 Powtórz", key=f"hist_repeat_{i}", use_container_width=True):
                        st.session_state["input_branza"] = branza_h
                        st.session_state["input_miasto"] = miasto_h
                        st.session_state["nav_page"] = PAGES[0]
                        st.rerun()

        if st.button("🗑️ Wyczyść historię zapytań", use_container_width=True):
            os.remove(HISTORY_FILE)
            if GCS_BUCKET:
                try:
                    _gcs_client().bucket(GCS_BUCKET).blob(HISTORY_FILE).delete()
                except Exception:
                    pass
            st.rerun()


# ── Strona: Ustawienia ─────────────────────────────────────────────────────────
elif page == "⚙️ Ustawienia":
    st.caption("Klucz API, integracja z CRM, trwałość danych oraz konfiguracja i test scoringu — wszystko w jednym miejscu.")

    with st.container(border=True):
        st.markdown("#### Klucz Google Places API")
        if env_api_key:
            st.markdown(pill_html("✓ Klucz ze zmiennej środowiskowej", "#1B8A5A", "#E6F7EF"), unsafe_allow_html=True)
        else:
            kcol1, kcol2 = st.columns([4, 1])
            with kcol1:
                st.text_input(
                    "Google Places API Key", type="password", placeholder="AIza...",
                    key="manual_api_key", label_visibility="collapsed",
                )
            with kcol2:
                if st.button("Zapisz", type="primary", use_container_width=True, key="save_api_key_btn"):
                    new_key = st.session_state["manual_api_key"].strip()
                    save_config({**app_config, "google_places_api_key": new_key})
                    # Verify the write actually landed (guards against a silent GCS write failure
                    # that would otherwise leave session_state and the persisted config out of sync).
                    verified = load_config().get("google_places_api_key", "") == new_key
                    print(f"[api_key debug] save attempted key_empty={not bool(new_key)} verified={verified}")
                    if verified:
                        st.session_state["manual_api_key"] = new_key
                        st.toast("✅ Klucz API zapisany")
                        st.rerun()
                    else:
                        st.toast("⚠️ Nie udało się potwierdzić zapisu klucza (błąd GCS?) — spróbuj ponownie")

            if st.session_state.get("manual_api_key"):
                st.markdown(pill_html("✓ Klucz zapisany i aktywny", "#1B8A5A", "#E6F7EF"), unsafe_allow_html=True)
            else:
                st.markdown(pill_html("⚠️ Klucz nie jest ustawiony", "#B7791F", "#FEF3D7"), unsafe_allow_html=True)
        st.caption("Klucz zapisywany w chmurze (GCS) — nie znika po odświeżeniu.")

    with st.container(border=True):
        st.markdown("#### Połączenie z CRM")
        st.caption(f"CRM_API_BASE_URL: {CRM_API_BASE_URL if CRM_API_BASE_URL else 'Nieskonfigurowany'}")
        if CRM_ENABLED:
            st.markdown(pill_html("✓ CRM połączony", "#1B8A5A", "#E6F7EF"), unsafe_allow_html=True)
        else:
            st.markdown(pill_html("✗ CRM nieskonfigurowany", "#D64545", "#FCEAEA"), unsafe_allow_html=True)
        if st.button("🔌 Test połączenia", key="crm_test_conn"):
            with st.spinner("Sprawdzam połączenie..."):
                ok, msg = crm_test_connection()
            st.toast(("✅ " if ok else "❌ ") + msg)

    with st.container(border=True):
        st.markdown("#### 📦 Trwałość danych i statystyki")
        gc1, gc2, gc3 = st.columns(3)
        gc1.metric("Leadów w bazie", len(load_master()))
        gc2.metric("Wykonanych zapytań", len(load_history()))
        gc3.metric("GCS bucket", GCS_BUCKET if GCS_BUCKET else "brak")
        st.caption("💡 **$200 free/miesiąc** ≈ 4 000 firm")

    current_weights = load_weights()

    with st.container(border=True):
      st.markdown("#### Konfiguracja scoringu")
      st.caption("Wagi poniżej są w pełni edytowalne — wartości startowe to tylko punkt wyjścia, zmień je lub usuń reguły dowolnie.")

      with st.expander("⚙️ Konfiguracja scoringu", expanded=True):
        st.markdown("**Status strony WWW** (dokładnie jeden z trzech stanów + osobny bonus za brak mobilności)")
        s1, s2, s3 = st.columns(3)
        with s1:
            w_brak_strony = st.number_input("Brak strony/domeny", min_value=0, value=int(current_weights["brak_strony"]), key="w_brak_strony")
        with s2:
            w_strona_nie_dziala = st.number_input("Jest domena, ale nie działa", min_value=0, value=int(current_weights["strona_nie_dziala"]), key="w_strona_nie_dziala")
        with s3:
            w_strona_dziala = st.number_input("Jest strona i działa", min_value=0, value=int(current_weights["strona_dziala"]), key="w_strona_dziala")
        w_niemobilna_bonus = st.number_input(
            "Bonus: niemobilna (brak <meta viewport>) — dolicza się do 'Jest strona i działa'",
            min_value=0, value=int(current_weights["niemobilna_bonus"]), key="w_niemobilna_bonus"
        )

        st.markdown("**Reguły punktowe — liczba opinii**")
        st.caption("Dla każdego leada stosowana jest reguła z najwyższym `min_count` <= liczba opinii. Dodawaj/usuwaj wiersze dowolnie.")
        df_opinie_rules = st.data_editor(
            pd.DataFrame(current_weights["reguly_opinii"]),
            num_rows="dynamic", use_container_width=True, key="editor_reguly_opinii",
            column_config={
                "min_count": st.column_config.NumberColumn("min_count", min_value=0, step=1, required=True),
                "points": st.column_config.NumberColumn("points", min_value=0, step=1, required=True),
            },
        )

        st.markdown("**Reguły punktowe — ocena Google**")
        st.caption("Dla każdego leada stosowana jest reguła z najwyższym `min_rating` <= ocena. Dodawaj/usuwaj wiersze dowolnie.")
        df_ocena_rules = st.data_editor(
            pd.DataFrame(current_weights["reguly_oceny"]),
            num_rows="dynamic", use_container_width=True, key="editor_reguly_oceny",
            column_config={
                "min_rating": st.column_config.NumberColumn("min_rating", min_value=0.0, max_value=5.0, step=0.1, required=True),
                "points": st.column_config.NumberColumn("points", min_value=0, step=1, required=True),
            },
        )

        live_weights = {
            "brak_strony": int(w_brak_strony),
            "strona_nie_dziala": int(w_strona_nie_dziala),
            "strona_dziala": int(w_strona_dziala),
            "niemobilna_bonus": int(w_niemobilna_bonus),
            "reguly_opinii": [
                {"min_count": int(r["min_count"]), "points": int(r["points"])}
                for r in df_opinie_rules.dropna().to_dict("records")
            ],
            "reguly_oceny": [
                {"min_rating": float(r["min_rating"]), "points": int(r["points"])}
                for r in df_ocena_rules.dropna().to_dict("records")
            ],
        }

        bcol1, bcol2 = st.columns(2)
        with bcol1:
            if st.button("💾 Zapisz ustawienia", type="primary", use_container_width=True, key="save_scoring_weights"):
                save_weights(live_weights)
                st.success("✅ Nowe wagi zostały zapisane i są teraz aktywne — będą stosowane we wszystkich kolejnych scrapingach oraz w teście scoringu poniżej.")
        with bcol2:
            if st.button("↩️ Przywróć wartości startowe", use_container_width=True, key="reset_scoring_weights"):
                save_weights(DEFAULT_WEIGHTS)
                st.success("✅ Przywrócono wartości startowe.")
                st.rerun()

      with st.expander("🧪 Test scoringu", expanded=False):
        st.markdown("Sprawdź jak `score_website()` ocenia konkretną firmę, na aktualnie zapisanych wagach.")

        tc1, tc2, tc3 = st.columns([2, 1, 1])
        with tc1:
            test_url = st.text_input("Adres WWW (puste = symulacja braku strony)", key="test_scoring_url", placeholder="https://przyklad.pl")
        with tc2:
            test_rating = st.number_input("Ocena (0-5)", min_value=0.0, max_value=5.0, value=4.0, step=0.1, key="test_scoring_rating")
        with tc3:
            test_reviews = st.number_input("Liczba opinii", min_value=0, value=10, step=1, key="test_scoring_reviews")

        if st.button("▶️ Uruchom scoring", type="primary", key="run_test_scoring"):
            with st.spinner("Sprawdzam stronę..."):
                score, status, breakdown = score_website(test_url.strip(), test_rating, test_reviews)

            st.markdown(f"## 🏆 {score} pkt")
            st.markdown(f"**Website status:** `{status}`")

            st.markdown("**Uzasadnienie punktacji:**")
            for line in format_breakdown(breakdown):
                st.markdown(f"- {line}")

            with st.expander("Surowy breakdown (JSON)", expanded=False):
                st.json(breakdown)
